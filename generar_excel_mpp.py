import os
import pandas as pd
import datetime

import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    from expand_gantt import ZONES
except ImportError:
    ZONES = []

ROWS = []
# Each zone is a Summary Task, and the 32 tasks are subtasks
start_date = datetime.date.today()

task_id = 1
for z in ZONES:
    zone_name = z['zone']
    ROWS.append({
        'ID': task_id,
        'Name': zone_name,
        'Outline Level': 1,
        'Duration': '32w',
        'Start': start_date.strftime('%Y-%m-%d'),
        'Finish': (start_date + datetime.timedelta(weeks=32)).strftime('%Y-%m-%d'),
    })
    task_id += 1
    
    current_date = start_date
    for t in z['tasks']:
        ROWS.append({
            'ID': task_id,
            'Name': t,
            'Outline Level': 2,
            'Duration': '1w',
            'Start': current_date.strftime('%Y-%m-%d'),
            'Finish': (current_date + datetime.timedelta(weeks=1)).strftime('%Y-%m-%d'),
        })
        current_date += datetime.timedelta(weeks=1)
        task_id += 1

df = pd.DataFrame(ROWS)
df.to_csv("Para_Microsoft_Project.csv", index=False)

# Excel Gantt visually
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Gantt"

# Headers
ws.cell(row=1, column=1, value="Zona/Actividad").font = Font(bold=True)
for w in range(1, 33):
    c = ws.cell(row=1, column=w+1, value=f"S{w}")
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center")
    ws.column_dimensions[openpyxl.utils.get_column_letter(w+1)].width = 4

ws.column_dimensions['A'].width = 30

row_idx = 2
for z in ZONES:
    c = ws.cell(row=row_idx, column=1, value=f"{z['emoji']} {z['zone']}")
    c.font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    for col in range(1, 34):
        ws.cell(row=row_idx, column=col).fill = fill
    row_idx += 1
    
    for w_idx, t in enumerate(z['tasks']):
        ws.cell(row=row_idx, column=1, value=t)
        # color the cell for the week
        t_cell = ws.cell(row=row_idx, column=w_idx+2)
        t_cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        row_idx += 1

wb.save("Cronograma_Sauces.xlsx")
print("Archivos generados: Cronograma_Sauces.xlsx y Para_Microsoft_Project.csv")
